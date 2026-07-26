#!/usr/bin/env python3
"""
Create an anonymized copy of the driving school's five source workbooks for a
public portfolio.

Three of the five tables share people, joined on Aday No (a clean numeric ID,
unlike Etstur's hotel names): Kursiyer Listesi, Kursiyer Genel Sınav ve Borç
Listesi, and Gelir Listesi. The fake identity for each person is built ONCE,
keyed by their real Aday No, then applied consistently across all three
tables. This avoids the Etstur problem entirely: there is no risk of the same
person getting two different fake identities just because their name is
formatted differently in different files (split ADI/SOYADI in one table,
combined ADI SOYADI in another).

Gider Listesi and "Diğer Gelir" Listesi carry no Aday No and are not
person-level records; they're masked independently.

T.C. Kimlik No and HESAP NO are not used anywhere in the report. Rather than
mask them, they're overwritten with an obvious placeholder so they can be
deleted entirely after loading, the same treatment used for the psychotechnical
project's T.C. Kimlik No and Sertifika No.

Money columns are scaled by a fixed factor, not randomized, so relative
amounts (and anything computed from them) stay meaningful.

Dates are left untouched throughout, including date of birth. A bare date
next to a fake name identifies no one.

Set the file paths below, then run.
"""

import os
import random

from openpyxl import Workbook, load_workbook

# ---- File paths (edit these) --------------------------------------------

TRAINEE_PATH   = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\İlt_ 2020-2025 MERKEZ VERİLERİ SON HALİ\Yıllık Kursiyer Listesi\Yıllık Kursiyer Listesi.xlsx"
EXPENSE_PATH   = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\İlt_ 2020-2025 MERKEZ VERİLERİ SON HALİ\Yıllık Gider Listesi\Gider Listesi.xlsx"
EXAM_DEBT_PATH = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\İlt_ 2020-2025 MERKEZ VERİLERİ SON HALİ\Yıllık Genel Sınav ve Borç Listesi\Kursiyer Genel Sınav ve Borç Listesi.xlsx"
INCOME_PATH    = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\İlt_ 2020-2025 MERKEZ VERİLERİ SON HALİ\Yıllık Gelir Listesi\Gelir Listesi.xlsx"
OTHER_INCOME_PATH = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\İlt_ 2020-2025 MERKEZ VERİLERİ SON HALİ\Yıllık Diğer Gelirler\Diğer Gelirler.xlsx"
OUT_DIR        = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\driving_mock"

SEED = 20260726  # fixed so re-runs produce the same anonymization
MONEY_FACTOR = 1.42  # every monetary value scales by this, same treatment as Etstur

# ---- Column names, as they appear in each file ---------------------------
# Edit these if your headers differ even slightly (extra space, different case).

ADAY_NO_COL = "ADAY NO"

# Kursiyer Listesi: split name
TRAINEE_FIRST_COL = "ADI"
TRAINEE_LAST_COL = "SOYADI"
TRAINEE_PHONE_COLS = ["1. TEL", "2. TEL"]

# Genel Sınav ve Borç Listesi + Gelir Listesi: combined name
COMBINED_NAME_COL = "ADI SOYADI"

# Columns that identify a person but never appear in the report. Filled with a
# placeholder here; delete them entirely after masking.
TC_NO_COL = "TC NO"                 # Gelir Listesi
HESAP_NO_COL = "HESAP NO"           # Gider Listesi, Diğer Gelir Listesi

# Money columns per table. "KALAN" appears twice in Genel Sınav ve Borç Listesi
# (remaining course debt, remaining exam fee) — both occurrences are scaled;
# the lookup below finds every column with a matching name, not just the first.
EXPENSE_MONEY_COLS = ["TUTAR"]
EXAM_DEBT_MONEY_COLS = ["TOPLAM BORÇ", "ÖDENEN", "KALAN", "SINAV HARÇ TOPLAMI", "ÖDENEN SINAV"]
INCOME_MONEY_COLS = ["TUTAR"]
OTHER_INCOME_MONEY_COLS = ["TUTAR"]

# --------------------------------------------------------------------------


def read_sheet(path):
    workbook = load_workbook(path)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])
    data = [list(r) for r in rows[1:]]
    return headers, data


def write_sheet(path, headers, data, sheet_title):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet.append(headers)
    for row in data:
        worksheet.append(row)
    workbook.save(path)


def all_indices(headers, name):
    """Every column index matching this header name, not just the first.
    Needed because Genel Sınav ve Borç Listesi repeats "KALAN"."""
    return [i for i, h in enumerate(headers) if h == name]


def scale_money(value, factor):
    if value is None or value == "":
        return value
    try:
        return round(float(value) * factor, 2)
    except (TypeError, ValueError):
        return value


def build_identity_map(trainee_headers, trainee_data, rng):
    """Build the Aday No -> fake number and Aday No -> fake name mappings
    ONCE, from the master trainee list. Every other table looks up by Aday No
    rather than building its own name-based mapping, so the same person gets
    the same fake identity everywhere regardless of how their name is
    formatted in that particular file."""
    aday_idx = trainee_headers.index(ADAY_NO_COL)

    real_aday_nos = []
    seen = set()
    for row in trainee_data:
        raw = row[aday_idx]
        if raw not in (None, "") and raw not in seen:
            seen.add(raw)
            real_aday_nos.append(raw)

    rng.shuffle(real_aday_nos)
    aday_map = {real: i + 1 for i, real in enumerate(real_aday_nos)}
    name_map = {real: f"Customer #{i + 1}" for i, real in enumerate(real_aday_nos)}
    return aday_map, name_map


def mask_trainee_list(headers, data, aday_map, name_map, rng):
    aday_idx = headers.index(ADAY_NO_COL)
    first_idx = headers.index(TRAINEE_FIRST_COL)
    last_idx = headers.index(TRAINEE_LAST_COL) if TRAINEE_LAST_COL in headers else None
    phone_idxs = [headers.index(c) for c in TRAINEE_PHONE_COLS if c in headers]

    for row in data:
        raw_aday = row[aday_idx]
        if raw_aday not in (None, ""):
            row[aday_idx] = aday_map[raw_aday]
            row[first_idx] = name_map[raw_aday]
            if last_idx is not None:
                row[last_idx] = ""  # full fake name lives in ADI, matching the psikoteknik precedent
        for i in phone_idxs:
            if row[i] not in (None, ""):
                row[i] = "".join(str(rng.randint(0, 9)) for _ in range(10))
    return data


def mask_person_linked_table(headers, data, aday_map, name_map):
    """For Genel Sınav ve Borç Listesi and Gelir Listesi: look up the fake
    identity by Aday No rather than building a new one, so it matches the
    trainee list exactly."""
    aday_idx = headers.index(ADAY_NO_COL)
    name_idx = headers.index(COMBINED_NAME_COL) if COMBINED_NAME_COL in headers else None

    for row in data:
        raw_aday = row[aday_idx]
        if raw_aday not in (None, ""):
            row[aday_idx] = aday_map.get(raw_aday, raw_aday)
            if name_idx is not None and raw_aday in name_map:
                row[name_idx] = name_map[raw_aday]
    return data


def apply_money_scaling(headers, data, money_col_names, factor):
    idxs = set()
    for name in money_col_names:
        idxs.update(all_indices(headers, name))
    for row in data:
        for i in idxs:
            row[i] = scale_money(row[i], factor)
    return data


def placeholder_column(headers, data, col_name):
    if col_name not in headers:
        return data
    idx = headers.index(col_name)
    for row in data:
        if row[idx] not in (None, ""):
            row[idx] = "REMOVE_ME"
    return data


def main():
    rng = random.Random(SEED)
    os.makedirs(OUT_DIR, exist_ok=True)

    # --- Master identity map, built once from the trainee list ---
    trainee_headers, trainee_data = read_sheet(TRAINEE_PATH)
    aday_map, name_map = build_identity_map(trainee_headers, trainee_data, rng)

    # --- Kursiyer Listesi ---
    trainee_data = mask_trainee_list(trainee_headers, trainee_data, aday_map, name_map, rng)
    trainee_out = os.path.join(OUT_DIR, os.path.basename(TRAINEE_PATH))
    write_sheet(trainee_out, trainee_headers, trainee_data, "Kursiyer Listesi")

    # --- Kursiyer Genel Sınav ve Borç Listesi ---
    exam_headers, exam_data = read_sheet(EXAM_DEBT_PATH)
    exam_data = mask_person_linked_table(exam_headers, exam_data, aday_map, name_map)
    exam_data = apply_money_scaling(exam_headers, exam_data, EXAM_DEBT_MONEY_COLS, MONEY_FACTOR)
    exam_out = os.path.join(OUT_DIR, os.path.basename(EXAM_DEBT_PATH))
    write_sheet(exam_out, exam_headers, exam_data, "Genel Sinav ve Borc Listesi")

    # --- Gelir Listesi ---
    income_headers, income_data = read_sheet(INCOME_PATH)
    income_data = mask_person_linked_table(income_headers, income_data, aday_map, name_map)
    income_data = apply_money_scaling(income_headers, income_data, INCOME_MONEY_COLS, MONEY_FACTOR)
    income_data = placeholder_column(income_headers, income_data, TC_NO_COL)
    income_out = os.path.join(OUT_DIR, os.path.basename(INCOME_PATH))
    write_sheet(income_out, income_headers, income_data, "Gelir Listesi")

    # --- Gider Listesi (no Aday No, not person-level) ---
    expense_headers, expense_data = read_sheet(EXPENSE_PATH)
    expense_data = apply_money_scaling(expense_headers, expense_data, EXPENSE_MONEY_COLS, MONEY_FACTOR)
    expense_data = placeholder_column(expense_headers, expense_data, HESAP_NO_COL)
    expense_out = os.path.join(OUT_DIR, os.path.basename(EXPENSE_PATH))
    write_sheet(expense_out, expense_headers, expense_data, "Gider Listesi")

    # --- "Diğer Gelir" Listesi (same shape as Gider Listesi) ---
    other_headers, other_data = read_sheet(OTHER_INCOME_PATH)
    other_data = apply_money_scaling(other_headers, other_data, OTHER_INCOME_MONEY_COLS, MONEY_FACTOR)
    other_data = placeholder_column(other_headers, other_data, HESAP_NO_COL)
    other_out = os.path.join(OUT_DIR, os.path.basename(OTHER_INCOME_PATH))
    write_sheet(other_out, other_headers, other_data, "Diger Gelir Listesi")

    print(f"{len(aday_map)} distinct candidates anonymized")
    print(f"  {trainee_out}")
    print(f"  {exam_out}")
    print(f"  {income_out}")
    print(f"  {expense_out}")
    print(f"  {other_out}")
    print("\nNext: delete the REMOVE_ME columns (TC NO, HESAP NO), then point")
    print("Power BI's data source settings at these files and refresh.")


if __name__ == "__main__":
    main()
